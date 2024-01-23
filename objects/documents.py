import openpyxl
from objects.zircon import Sample, Grain


class SampleSheet:
    def __init__(self, file):
        self.file = file

    def read_samples(self):
        file = self.file
        samples = []
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        for i in range(1, sheet.max_column + 1, 2):
            sample_name = sheet.cell(row=1, column=i).value
            grains = []
            for row in range(2, sheet.max_row + 1):
                age = sheet.cell(row=row, column=i).value
                uncertainty = sheet.cell(row=row, column=i + 1).value
                if age is not None and uncertainty is not None:
                    grains.append(Grain(age, uncertainty))
            sample = Sample(sample_name, grains)
            samples.append(sample)
        samples.reverse()
        return samples

    def create_mixed_sample(self):
        samples = self.read_samples()

        # Find the sample with the maximum number of grains
        longest_sample = max(samples, key=lambda sample: len(sample.grains))

        target_length = len(longest_sample.grains)
        sample_name = "Mixed Sample"
        sample_grains = []

        for i in range(target_length):
            age_sum = 0
            uncertainty_sum = 0
            samples_with_grains = 0

            for sample in samples:
                if i < len(sample.grains) and sample.grains[i] is not None:
                    age_sum += sample.grains[i].age
                    uncertainty_sum += sample.grains[i].uncertainty
                    samples_with_grains += 1

            if samples_with_grains > 0:
                age_average = age_sum / samples_with_grains
                uncertainty_average = uncertainty_sum / samples_with_grains
                sample_grains.append(Grain(age_average, uncertainty_average))
            else:
                # Handle the case where no sample has grains at the current index
                sample_grains.append(None)

        mixed_sample = Sample(sample_name, sample_grains)
        return mixed_sample

    def __is_sample_sheet(self):
        # TODO: check if the file is formatted correctly, and then work this into the read_samples function
        if True:
            return True
        elif False:
            return False
