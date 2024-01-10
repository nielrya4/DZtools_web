import openpyxl
import pandas as pd
import numpy as np
import os
import app
from utils import format, measures
import pickle
from werkzeug.utils import secure_filename

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def read_excel(file):
    try:
        wb = openpyxl.load_workbook(file, read_only=False)
        sheet = wb.active

        num_cols = 0
        for col in sheet.iter_cols(values_only=True):
            num_cols += 1

        all_data = []
        for col in range(1, num_cols + 1, 2):
            if sheet.cell(row=1, column=col).value is not None:
                header = f"{sheet.cell(row=1, column=col).value}"
                data = format.trim_none([value for value in sheet.iter_rows(min_col=col,
                                                                            max_col=col,
                                                                            min_row=2,
                                                                            values_only=True)])

                bandwidth = format.trim_none([value for value in sheet.iter_rows(min_col=col + 1,
                                                                                 max_col=col + 1,
                                                                                 min_row=2,
                                                                                 values_only=True)])
                data_set = [header, data, bandwidth]
                all_data.append(data_set)

    except Exception as e:
        print(f"{e}")
        raise ValueError(f"Error parsing Excel file: {e}")

    return all_data


def generate_matrix(y_value_arrays, row_labels=None, col_labels=None, matrix_type="similarity"):
    num_data_sets = len(y_value_arrays)
    matrix = np.zeros((num_data_sets, num_data_sets))

    if matrix_type == "similarity":
        for i, y1 in enumerate(y_value_arrays):
            for j, y2 in enumerate(y_value_arrays):
                similarity_score = measures.similarity_test(y1, y2)
                matrix[i, j] = similarity_score
    elif matrix_type == "likeness":
        for i, y1 in enumerate(y_value_arrays):
            for j, y2 in enumerate(y_value_arrays):
                likeness_score = measures.likeness_test(y1, y2)
                matrix[i, j] = likeness_score
    elif matrix_type == "ks":
        for i, y1 in enumerate(y_value_arrays):
            for j, y2 in enumerate(y_value_arrays):
                ks_score = measures.ks_test(y1, y2)
                matrix[i, j] = ks_score
    elif matrix_type == "kuiper":
        for i, y1 in enumerate(y_value_arrays):
            for j, y2 in enumerate(y_value_arrays):
                kuiper_score = measures.kuiper_test(y1, y2)
                matrix[i, j] = kuiper_score
    elif matrix_type == "cross_correlation":
        for i, y1 in enumerate(y_value_arrays):
            for j, y2 in enumerate(y_value_arrays):
                cross_correlation_score = measures.cross_correlation_test(y1, y2)
                matrix[i, j] = cross_correlation_score

    # Create a DataFrame with the normalized similarity scores and labels
    if row_labels is None:
        row_labels = [f'Data {i+1}' for i in range(num_data_sets)]
    if col_labels is None:
        col_labels = [f'Data {i+1}' for i in range(num_data_sets)]

    df = pd.DataFrame(matrix, columns=col_labels, index=row_labels)
    return df


def save_data_to_file(data, filepath):
    with open(filepath, 'wb') as file:
        pickle.dump(data, file)


def read_data_from_file(filepath):
    with open(filepath, 'rb') as file:
        return pickle.load(file)


def get_extension(file):
    # Gets the extension of a file. I.E. ".xlsx" or ".txt"
    if file.filename == '':
        return None
    filename = secure_filename(file.filename)
    return filename.rsplit('.', 1)[1].lower() if '.' in filename else None


def upload_file(file, session_key):
    # Upload a file to the specified folder with the session key in the filename.
    if file:
        filename = f"{session_key}_{secure_filename(file.filename)}"
        file_path = os.path.join(app.UPLOAD_FOLDER, filename)
        file.save(file_path)
        return file_path
    return None


def file_in_folder(folder_path):
    # Check if there is any file in the given folder.
    files = os.listdir(folder_path)
    return any(os.path.isfile(os.path.join(folder_path, filename)) for filename in files)


def get_latest_file(folder_path):
    # Get the latest file in the given folder based on modification time.
    files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
    if not files:
        return None
    latest_file = max(files, key=lambda f: os.path.getmtime(os.path.join(folder_path, f)))
    return os.path.join(folder_path, latest_file)
