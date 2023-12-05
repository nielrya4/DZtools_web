import openpyxl
import numpy as np

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def read_excel(file):
    try:
        wb = openpyxl.load_workbook(file, read_only=False)
        sheet = wb.active

        num_cols = 0
        for col in sheet.iter_cols(values_only=True):
            num_cols += 1

        all_data = []
        for col in range(1, num_cols + 1, 2):
            if sheet.cell(row=1, column=col).value != None:
                header = f"{sheet.cell(row=1, column=col).value}"
                data = [value for value in sheet.iter_rows(min_col=col, max_col=col, min_row=2, values_only=True)]
                sigma = [value for value in sheet.iter_rows(min_col=col + 1, max_col=col + 1, min_row=2, values_only=True)]
                data_set = [header, data, sigma]
                all_data.append(data_set)

    except Exception as e:
        print(f"{e}")
        raise ValueError(f"Error parsing Excel file: {e}")

    return all_data



