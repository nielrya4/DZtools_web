import os
import openpyxl
from flask import render_template, request, jsonify, session, redirect
from openpyxl import Workbook
import app as APP
from lib.objects.documents import SampleSheet
from lib.utils import dz_script, files




def register(app):
    @app.route('/data_editor/', methods=['GET', 'POST'])
    def data_editor():
        session_key = session.get('SECRET_KEY', APP.SECRET_KEY)
        spreadsheet_data = []
        output_items = []

        kde_graph = request.form.get('kde_graph') == "true"
        pdp_graph = request.form.get('pdp_graph') == "true"
        cdf_graph = request.form.get('cdf_graph') == "true"
        mds_graph = request.form.get('mds_graph') == "true"

        similarity_matrix = request.form.get('similarity_matrix') == "true"
        dissimilarity_matrix = request.form.get('dissimilarity_matrix') == "true"
        likeness_matrix = request.form.get('likeness_matrix') == "true"
        ks_matrix = request.form.get('ks_matrix') == "true"
        kuiper_matrix = request.form.get('kuiper_matrix') == "true"
        cross_correlation_matrix = request.form.get('cross_correlation_matrix') == "true"

        # First, make sure we have a file in the session no matter what, and make sure we have some spreadsheet_data
        if "last_uploaded_file" in session:
            last_uploaded_file = session.get("last_uploaded_file")
            file = os.path.join(APP.UPLOAD_FOLDER, last_uploaded_file)
            workbook = openpyxl.load_workbook(file)
            sheet = workbook.active
            max_rows = sheet.max_row
            max_cols = sheet.max_column
            for row in range(1, max_rows + 1):
                row_data = []
                for col in range(1, max_cols + 1):
                    cell_value = sheet.cell(row=row, column=col).value
                    # Replace empty cell values with None
                    cell_value = cell_value if cell_value is not None else None
                    row_data.append(cell_value)
                spreadsheet_data.append(row_data)
        else:
            spreadsheet_data = [[None for _ in range(6)] for _ in range(6)]
            wb = Workbook()
            ws = wb.active
            for col_idx, column in enumerate(spreadsheet_data, start=1):
                for row_idx, value in enumerate(column, start=1):
                    ws.cell(row=col_idx, column=row_idx, value=value)
            filename = 'spreadsheet.xlsx'
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_key}_{filename}")
            wb.save(str(filepath))
            session['last_uploaded_file'] = os.path.basename(filepath)

        #Now see if we also have a script file uploaded...
        if "last_uploaded_script_file" not in session:
            filename = 'script.dzs'
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_key}_{filename}")
            session['last_uploaded_script_file'] = os.path.basename(filepath)
            with open(filepath, 'w') as f:
                f.write("load $all\npurge $all\n")

        #Now that we for sure have a file and spreadsheet data, generate output data
        if session.get("last_uploaded_file") is not None and session.get("last_uploaded_script_file") is not None:

            file = os.path.join(APP.UPLOAD_FOLDER, session.get("last_uploaded_script_file"))
            kde_bandwidth = int(session.get('kde_bandwidth', 10))
            stacked = request.args.get('stacked') == 'true'

            kde_graph = request.form.get('kde_graph') == "true"
            pdp_graph = request.form.get('pdp_graph') == "true"
            cdf_graph = request.form.get('cdf_graph') == "true"
            mds_graph = request.form.get('mds_graph') == "true"
            similarity_matrix = request.form.get('similarity_matrix') == "true"
            dissimilarity_matrix = request.form.get('dissimilarity_matrix') == "true"
            likeness_matrix = request.form.get('likeness_matrix') == "true"
            ks_matrix = request.form.get('ks_matrix') == "true"
            kuiper_matrix = request.form.get('kuiper_matrix') == "true"
            cross_correlation_matrix = request.form.get('cross_correlation_matrix') == "true"

            clear_script_file(file)
            update_script_file(file, "load $all")

            if kde_graph:
                update_script_file(file, "stacked" if stacked else "unstacked")
                update_script_file(file, "kde")
            if pdp_graph:
                update_script_file(file, "stacked" if stacked else "unstacked")
                update_script_file(file, "pdp")
            if cdf_graph:
                update_script_file(file, "cdf")
            if mds_graph:
                update_script_file(file, "mds")
            if similarity_matrix:
                update_script_file(file, "sim")
            if dissimilarity_matrix:
                update_script_file(file, "dis")
            if likeness_matrix:
                update_script_file(file, "lik")
            if ks_matrix:
                update_script_file(file, "ks")
            if kuiper_matrix:
                update_script_file(file, "kpr")
            if cross_correlation_matrix:
                update_script_file(file, "ccr")

            update_script_file(file, "purge $all")

            last_uploaded_file = session.get("last_uploaded_file")
            file = os.path.join(APP.UPLOAD_FOLDER, last_uploaded_file)
            sample_sheet = SampleSheet(file)
            samples = sample_sheet.read_samples()

            last_uploaded_script_file = session.get("last_uploaded_script_file")
            code_file = os.path.join(APP.UPLOAD_FOLDER, last_uploaded_script_file)
            output_items = dz_script.run(code_file, samples)

            for i, output_item in enumerate(output_items):
                output_item = f"<div id='output{i}'>{output_item}</div>"
        return render_template('data_editor/data_editor.html',
                               spreadsheet_data=spreadsheet_data,
                               output_items=output_items,
                               kde_graph=kde_graph,
                               pdp_graph=pdp_graph,
                               cdf_graph=cdf_graph,
                               mds_graph=mds_graph,
                               similarity_matrix=similarity_matrix,
                               dissimilarity_matrix=dissimilarity_matrix,
                               likeness_matrix=likeness_matrix,
                               ks_matrix=ks_matrix,
                               kuiper_matrix=kuiper_matrix,
                               cross_correlation_matrix=cross_correlation_matrix
                               )

    @app.route('/json/save', methods=['POST'])
    def json_save():
        data = request.get_json()['data']
        session_key = session.get('SECRET_KEY', APP.SECRET_KEY)

        # Create a new Workbook
        wb = Workbook()
        ws = wb.active

        # Write the data to the worksheet
        for col_idx, column in enumerate(data, start=1):
            for row_idx, value in enumerate(column, start=1):
                ws.cell(row=col_idx, column=row_idx, value=value)

        # Save the workbook to a file in the uploads folder
        if session.get('last_uploaded_file', 0) == 0:
            filename = 'spreadsheet.xlsx'
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"{session_key}_{filename}")
            session['last_uploaded_file'] = os.path.basename(filepath)
            print(session["last_uploaded_file"])
        else:
            filepath = os.path.join(APP.UPLOAD_FOLDER, session['last_uploaded_file'])
            print(filepath)
        wb.save(str(filepath))
        return jsonify({"result": "ok", "filename": filepath})

    @app.route('/data_editor/generate_outputs', methods=['POST'])
    def insert_item():
        file = os.path.join(APP.UPLOAD_FOLDER, session.get("last_uploaded_script_file"))
        kde_bandwidth = int(session.get('kde_bandwidth', 10))
        stacked = request.args.get('stacked') == 'true'

        kde_graph = request.form.get('kde_graph') == "true"
        pdp_graph = request.form.get('pdp_graph') == "true"
        cdf_graph = request.form.get('cdf_graph') == "true"
        mds_graph = request.form.get('mds_graph') == "true"

        similarity_matrix = request.form.get('similarity_matrix') == "true"
        dissimilarity_matrix = request.form.get('dissimilarity_matrix') == "true"
        likeness_matrix = request.form.get('likeness_matrix') == "true"
        ks_matrix = request.form.get('ks_matrix') == "true"
        kuiper_matrix = request.form.get('kuiper_matrix') == "true"
        cross_correlation_matrix = request.form.get('cross_correlation_matrix') == "true"

        clear_script_file(file)

        if kde_graph:
            update_script_file(file, "stacked" if stacked else "unstacked")
            update_script_file(file, "kde")
        if pdp_graph:
            update_script_file(file, "stacked" if stacked else "unstacked")
            update_script_file(file, "pdp")
        if cdf_graph:
            update_script_file(file, "cdf")
        if mds_graph:
            update_script_file(file, "mds")
        if similarity_matrix:
            update_script_file(file, "sim")
        if dissimilarity_matrix:
            update_script_file(file, "dis")
        if likeness_matrix:
            update_script_file(file, "lik")
        if ks_matrix:
            update_script_file(file, "ks")
        if kuiper_matrix:
            update_script_file(file, "kpr")
        if cross_correlation_matrix:
            update_script_file(file, "ccr")

        update_script_file(file, "purge $all")

        return redirect(request.referrer)

    def update_script_file(filepath, line):
        with open(filepath, 'a') as file:
            file.write(line + '\n')

    def clear_script_file(filepath):
        with open(filepath, 'a') as file:
            file.truncate(0)
