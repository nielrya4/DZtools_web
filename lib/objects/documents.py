import openpyxl
from lib.objects.zircon import Sample, Grain
from lib.utils import sample_utils


class SampleSheet:
    def __init__(self, file):
        self.file = file

    def read_samples(self):
        file = self.file
        samples = []
        workbook = openpyxl.load_workbook(file)
        sheet = workbook.active
        for i in range(1, sheet.max_column + 1, 2):
            sample_name = sheet.cell(row=1, column=i).value
            grains = []
            for row in range(2, sheet.max_row + 1):
                age = sheet.cell(row=row, column=i).value
                uncertainty = sheet.cell(row=row, column=i + 1).value
                if age is not None and uncertainty is not None:
                    grains.append(Grain(float(age), float(uncertainty)))
            sample = Sample(sample_name, grains)
            samples.append(sample)
        samples.reverse()
        return samples

    def create_mean_sample(self):
        samples = self.read_samples()
        return sample_utils.create_mean_sample(samples)

    def create_mixed_sample(self):
        samples = self.read_samples()
        return sample_utils.create_mixed_sample(samples)

    def __is_sample_sheet(self):
        # TODO: check if the file is formatted correctly, and then work this into the read_samples function
        if True:
            return True
        elif False:
            return False


class Template:
    def __init__(self, file):
        self.file = file
        self.lines = file.readlines()

    def execute(self):
        pass
